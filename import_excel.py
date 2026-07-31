"""
Script d'import : migre les données depuis un fichier Excel vers UMLM.

USAGE :
    python manage.py shell < import_excel.py

AVANT DE LANCER :
1. Placez votre fichier Excel à la racine du projet et ajustez EXCEL_PATH
   ci-dessous.
2. Adaptez les noms de colonnes (COL_*) pour qu'ils correspondent EXACTEMENT
   aux en-têtes de vos feuilles Excel (respecter la casse).
3. Testez d'abord sur une copie de votre fichier avec seulement 2-3 lignes,
   pour vérifier que le mapping est correct.

Ce script est idempotent pour les Rôles et Modules (utilise get_or_create),
mais PAS pour les Projets/Utilisateurs si vous n'avez pas de champ unique
adapté — relancer le script plusieurs fois peut créer des doublons pour ces
deux-là. Vérifiez toujours le résultat dans l'interface après un import.
"""
import openpyxl
from core.models import Role, User, Project, Module

# ==========================================================================
# 1. CONFIGURATION — à adapter à votre fichier
# ==========================================================================

EXCEL_PATH = "donnees_excel.xlsx"

# --- Feuille "Rôles" ---
SHEET_ROLES = "Roles"
COL_ROLE_NAME = "Nom"
COL_ROLE_DROITS = "Droits"   # texte libre, ex: "Gestion complète"

# --- Feuille "Modules" ---
SHEET_MODULES = "Modules"
COL_MODULE_NAME = "Nom"
COL_MODULE_GITHUB_URL = "URL GitHub"
COL_MODULE_BRANCH = "Branche"
COL_MODULE_CRITICALITY = "Criticité"   # doit valoir: faible / moyen / critique

# --- Feuille "Projets" ---
SHEET_PROJECTS = "Projets"
COL_PROJECT_NAME = "Nom"
COL_PROJECT_LOCAL_PATH = "Chemin local"
COL_PROJECT_BRANCH = "Branche"

# --- Feuille "Utilisateurs" ---
SHEET_USERS = "Utilisateurs"
COL_USER_NAME = "Nom"
COL_USER_EMAIL = "Email"
COL_USER_ROLE = "Rôle"   # doit correspondre exactement au Nom d'un Rôle importé plus haut


# ==========================================================================
# 2. FONCTIONS D'IMPORT — pas besoin de modifier en dessous de cette ligne
# ==========================================================================

def _rows(sheet):
    """Retourne les lignes de données sous forme de dict {en-tête: valeur}, en sautant les lignes vides."""
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        yield dict(zip(headers, row))


def import_roles(wb):
    if SHEET_ROLES not in wb.sheetnames:
        print(f"  (feuille '{SHEET_ROLES}' introuvable — étape sautée)")
        return
    sheet = wb[SHEET_ROLES]
    count = 0
    for row in _rows(sheet):
        name = row.get(COL_ROLE_NAME)
        if not name:
            continue
        role, created = Role.objects.get_or_create(name=str(name).strip())
        # Les droits texte libre ne sont pas remappés automatiquement vers
        # les cases à cocher — cochez-les manuellement dans l'interface
        # après l'import, ou adaptez cette fonction si le texte suit un
        # format prévisible.
        count += 1
    print(f"  {count} rôle(s) traité(s).")


def import_modules(wb):
    if SHEET_MODULES not in wb.sheetnames:
        print(f"  (feuille '{SHEET_MODULES}' introuvable — étape sautée)")
        return
    sheet = wb[SHEET_MODULES]
    count = 0
    for row in _rows(sheet):
        name = row.get(COL_MODULE_NAME)
        if not name:
            continue
        Module.objects.get_or_create(
            name=str(name).strip(),
            defaults={
                "github_url": row.get(COL_MODULE_GITHUB_URL, "") or "",
                "reference_branch": row.get(COL_MODULE_BRANCH, "main") or "main",
                "criticality": (row.get(COL_MODULE_CRITICALITY, "moyen") or "moyen").strip().lower(),
            },
        )
        count += 1
    print(f"  {count} module(s) traité(s).")


def import_projects(wb):
    if SHEET_PROJECTS not in wb.sheetnames:
        print(f"  (feuille '{SHEET_PROJECTS}' introuvable — étape sautée)")
        return
    sheet = wb[SHEET_PROJECTS]
    count = 0
    for row in _rows(sheet):
        name = row.get(COL_PROJECT_NAME)
        if not name:
            continue
        Project.objects.get_or_create(
            name=str(name).strip(),
            defaults={
                "local_path": row.get(COL_PROJECT_LOCAL_PATH, "") or "",
                "reference_branch": row.get(COL_PROJECT_BRANCH, "main") or "main",
            },
        )
        count += 1
    print(f"  {count} projet(s) traité(s).")


def import_users(wb):
    if SHEET_USERS not in wb.sheetnames:
        print(f"  (feuille '{SHEET_USERS}' introuvable — étape sautée)")
        return
    sheet = wb[SHEET_USERS]
    count, skipped = 0, 0
    for row in _rows(sheet):
        name = row.get(COL_USER_NAME)
        email = row.get(COL_USER_EMAIL)
        role_name = row.get(COL_USER_ROLE)
        if not name or not email:
            continue

        role = Role.objects.filter(name=str(role_name).strip()).first() if role_name else None
        if role is None:
            print(f"    ! rôle '{role_name}' introuvable pour l'utilisateur '{name}' — ligne ignorée")
            skipped += 1
            continue

        User.objects.get_or_create(
            email=str(email).strip(),
            defaults={"name": str(name).strip(), "role": role},
        )
        count += 1
    print(f"  {count} utilisateur(s) traité(s), {skipped} ignoré(s).")


# ==========================================================================
# 3. EXÉCUTION
# ==========================================================================

print(f"Ouverture de {EXCEL_PATH}...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

print("\nImport des rôles...")
import_roles(wb)

print("\nImport des modules...")
import_modules(wb)

print("\nImport des projets...")
import_projects(wb)

print("\nImport des utilisateurs...")
import_users(wb)

print("\nTerminé. Vérifiez le résultat dans l'interface UMLM.")
